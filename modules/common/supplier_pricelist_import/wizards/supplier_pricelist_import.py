# -*- coding: utf-8 -*-
import base64
import logging
from io import BytesIO
from openpyxl import load_workbook

from odoo import fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class SupplierPricelistImport(models.TransientModel):
    _name = "supplier.pricelist.import"
    _description = "Import Supplier Pricelist"

    ###########################################################################
    # Default and compute methods.
    ###########################################################################

    ###########################################################################
    # Fields
    ###########################################################################
    company = fields.Many2one("res.company", default=lambda self: self.env.company.id)
    company_parent = fields.Many2one("res.company", related="company.parent_id")

    import_for_all = fields.Boolean(string="Import for all companies",
                                    default=lambda self: not self.env.company.parent_id)

    update_normal_buy_price = fields.Boolean(string="Update Normal Buy Price")

    supplier = fields.Many2one("res.partner", string="Supplier", required=True)
    currency = fields.Many2one("res.currency", default=lambda self: self.env.company.currency_id.id, required=True)
    start = fields.Date("Date Start", help="Leave empty for update to use today's date")
    file = fields.Binary("File", required=True,
                         help=("Select the CSV file to import. The format should be:\n"
                               "Product-Code,Unit-Price,Min-QTY,Vendor-Code,Lead-Time,Is Preferred Supplier\n"
                               "e.g. \"ABC01\",1.50,75,\"XYABC-484\",5"))
    notes = fields.Text("Notes", readonly=True)

    ###########################################################################
    # Model Methods
    ###########################################################################
    def function_update_normal_buy_price(self, product, unit_price):
        """Updates normal buy price using JSON field instead of ir.property"""
        prices = product.normal_buy_price_json or {}

        if self.import_for_all:
            # update for all companies
            for company in self.env['res.company'].search([]):
                prices[f"company_{company.id}"] = unit_price
        else:
            company_id = self.company.id if self.company else self.env.company.id
            prices[f"company_{company_id}"] = unit_price

        product.normal_buy_price_json = prices

    def run_after_import_actions(self, product):
        pass

    def button_import(self):
        info_model = self.env["product.supplierinfo"]
        product_model = self.env["product.product"].with_context(exact_match=True)

        notes = ""
        error_count = 0
        process_count = 0

        # Decode and load Excel workbook
        try:
            file_data = base64.b64decode(self.file)
            workbook = load_workbook(filename=BytesIO(file_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(f"Unable to read XLSX file. Please upload a valid Excel file.\nError: {str(e)}")

        # Iterate through rows (skip header)
        for line_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if line_no == 1:
                continue  # skip header

            product_code, unit_price, min_qty, vendor_code, lead_time, preferred = row[:6] if row else (None,) * 6

            # Skip empty lines
            if not product_code:
                continue

            product = product_model.search([("default_code", "=", str(product_code).strip())])
            if not product:
                _logger.warning(f"Couldn't find product={product_code}")
                notes += f"** Line {line_no} - Product not found: {product_code}\n"
                error_count += 1
                continue

            # Validate and convert unit price
            try:
                unit_price = float(unit_price)
            except (TypeError, ValueError):
                _logger.warning(f"Invalid price={unit_price}")
                notes += f"** Line {line_no} - Invalid price: {unit_price}\n"
                error_count += 1
                continue

            # Convert other values
            min_qty = float(min_qty or 0)
            vendor_code = str(vendor_code or product_code)
            lead_time = int(lead_time or 1)
            new_preferred = str(preferred).strip().lower() in ("y", "yes", "true", "t", "1")

            if len(product) > 1:
                raise UserError(
                    f"There are {len(product)} products for Reference: {product[0].default_code}.\n"
                    f"Please keep just one of them and archive the rest."
                )

            # Find existing supplier info
            same_suppliers = info_model.search([
                ("partner_id", "=", self.supplier.id),
                ("product_tmpl_id", "=", product.product_tmpl_id.id)
            ])

            date_start = self.start or fields.Date.context_today(self)
            if same_suppliers:
                to_update = same_suppliers.filtered(lambda r: not r.date_end or r.date_end >= date_start)
                if to_update:
                    to_update.write({"date_end": date_start})

            # Preferred supplier sequence logic
            if new_preferred:
                sequence = 1
                for suppl_info in product.seller_ids:
                    suppl_info.sequence = suppl_info.sequence + 1
            else:
                seq_list = product.seller_ids.mapped("sequence")
                sequence = (max(seq_list) + 1) if seq_list else 1

            # Handle company
            company_id = None
            if self.company and not self.import_for_all:
                company_id = self.company.id

            # Create supplierinfo record
            info_model.create({
                "partner_id": self.supplier.id,
                "product_id": product.id,
                "product_tmpl_id": product.product_tmpl_id.id,
                "price": unit_price,
                "currency_id": self.currency.id,
                "min_qty": min_qty,
                "product_code": vendor_code,
                "delay": lead_time,
                "date_start": date_start,
                "sequence": sequence,
                "company_id": company_id,
            })

            # Update normal buy price if required (re-enable later if needed)
            # if self.update_normal_buy_price:
            #     self.function_update_normal_buy_price(product=product, unit_price=unit_price)

            self.run_after_import_actions(product=product)
            process_count += 1

        self.notes = f"{notes}\nProcessed: {process_count}, Errors: {error_count}"

        return {
            "name": self._description,
            "view_mode": "form",
            "res_model": self._name,
            "res_id": self.id,
            "type": "ir.actions.act_window",
            "target": "new",
        }

