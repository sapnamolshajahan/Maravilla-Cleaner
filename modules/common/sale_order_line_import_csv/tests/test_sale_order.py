# -*- coding: utf-8 -*-
import logging

from io import BytesIO
from odoo.tests import common, tagged
from odoo.exceptions import UserError
import base64
from openpyxl import Workbook

_logger = logging.getLogger(__name__)

@tagged("common", "sale_order_line_import_csv")
class TestSaleImportCsv(common.TransactionCase):
    """Class to test operation related sale.order import csv workflow"""

    def setUp(self):
        """
        Extend setUp functionality to create data used in tests
        """
        super().setUp()
        self.product_1 = self.env['product.product'].create({
            'name': 'Test Product1',
            'default_code': 'TEST_PRODUCT_1',
            'list_price': 100.0,
        })
        self.product_2 = self.env['product.product'].create({
            'name': 'Test Product2',
            'default_code': 'TEST_PRODUCT_2',
            'list_price': 100.0,
        })

        # Create a test warehouse
        self.warehouse = self.env['stock.warehouse'].create({
            'name': 'Test Warehouse',
            'code': 'TESTWH',
        })
        self.sale_order = self.env['sale.order'].create({
            'partner_id': self.env.ref('base.res_partner_1').id,
            'order_line': [(0, 0, {
                    'product_id': self.env.ref('product.product_product_12').id,
                    'product_uom_qty': 1,
                    'price_unit': 100,
                })
            ],
        })


    def test_action_import_sale_order_lines_draft(self):
        """Test that the wizard is created successfully when the sale order is in draft state."""
        self.sale_order.state = 'draft'
        action = self.sale_order.with_context(active_id=self.sale_order.id).action_import_sale_order_lines()
        self.assertEqual(action['res_model'], 'sale.order.line.import', "Incorrect res_model in action.")
        self.assertEqual(action['type'], 'ir.actions.act_window', "Incorrect action type.")
        self.assertEqual(action['target'], 'new', "The wizard should open in a new window.")
        wizard = self.env['sale.order.line.import'].browse(action['res_id'])
        self.assertEqual(wizard.order_id.id, self.sale_order.id, "Wizard created with incorrect sale order.")

    def test_action_import_sale_order_lines_non_draft(self):
        """Test that a UserError is raised when the sale order is not in draft state."""
        self.sale_order.state = "sale"
        with self.assertRaises(UserError):
            self.sale_order.action_import_sale_order_lines()

    def _create_test_excel_file(self):
        """ Helper to create a test Excel file in memory """
        workbook = Workbook()
        sheet = workbook.active
        # Add header row if needed
        sheet.append(["product_ref", "quantity", "discount", "warehouse"])
        # Add test rows
        sheet.append(["TEST_PRODUCT_1", 5, 10, "Test Warehouse"])
        sheet.append(["TEST_PRODUCT_2", 3, None, "Test Warehouse"])
        # Save the workbook to a BytesIO stream
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)
        return excel_stream

    def test_import_csv_success(self):
        """Test successful import of sale order lines from a valid CSV."""
        excel_file = self._create_test_excel_file()
        encoded_file = base64.b64encode(excel_file.getvalue())
        wizard = self.env['sale.order.line.import'].create({
            'order_id': self.sale_order.id,
            'csv_file': encoded_file,
            'csv_delimiter': ',',
        })
        wizard.button_import_csv()
        order_lines = self.sale_order.order_line
        self.assertEqual(len(order_lines), 3, "Expected 3 sale order lines after import")
        product_line1 = order_lines.filtered(lambda l: l.product_id.default_code == "TEST_PRODUCT_1")
        self.assertTrue(product_line1, "Product TEST_PRODUCT_1 was not imported")
        self.assertEqual(product_line1.product_uom_qty, 5, "Incorrect quantity for TEST_PRODUCT_1")
        self.assertEqual(product_line1.discount, 10, "Incorrect discount for TEST_PRODUCT_1")
        # Verify second product
        product_line2 = order_lines.filtered(lambda l: l.product_id.default_code == "TEST_PRODUCT_2")
        self.assertTrue(product_line2, "Product TEST_PRODUCT_2 was not imported")
        self.assertEqual(product_line2.product_uom_qty, 3, "Incorrect quantity for TEST_PRODUCT_2")
        self.assertEqual(product_line2.discount, 0, "Discount should be 0 for TEST_PRODUCT_2")

    def test_import_csv_missing_product(self):
        """Test import with missing product reference"""
        excel_file = self._create_test_excel_file()
        # Modify the test data to include a non-existent product reference
        excel_file.seek(0)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["product_ref", "quantity", "discount", "warehouse"])
        sheet.append(["NON_EXISTENT_PRODUCT", 5, 10, "Test Warehouse"])
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)
        encoded_file = base64.b64encode(excel_stream.getvalue())
        wizard = self.env['sale.order.line.import'].create({
            'order_id': self.sale_order.id,
            'csv_file': encoded_file,
            'csv_delimiter': ',',
            'ignore_errors': False
        })
        with self.assertRaises(UserError):
            wizard.button_import_csv()

    def test_import_csv_invalid_quantity(self):
        """Test import with invalid quantity"""
        excel_file = self._create_test_excel_file()
        # Modify the test data to include an invalid quantity (non-numeric)
        excel_file.seek(0)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["product_ref", "quantity", "discount", "warehouse"])
        sheet.append(["TEST_PRODUCT_1", "INVALID_QUANTITY", 10, "Test Warehouse"])
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)
        encoded_file = base64.b64encode(excel_stream.getvalue())
        wizard = self.env['sale.order.line.import'].create({
            'order_id': self.sale_order.id,
            'csv_file': encoded_file,
            'csv_delimiter': ',',
            'ignore_errors': False
        })
        with self.assertRaises(UserError):
            wizard.button_import_csv()

    def test_import_csv_invalid_warehouse(self):
        """Test import with invalid quantity"""
        excel_file = self._create_test_excel_file()
        # Modify the test data to include an invalid quantity (non-numeric)
        excel_file.seek(0)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["product_ref", "quantity", "discount", "warehouse"])
        sheet.append(["TEST_PRODUCT_1", 4, 10, "Test #"])
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)
        encoded_file = base64.b64encode(excel_stream.getvalue())
        wizard = self.env['sale.order.line.import'].create({
            'order_id': self.sale_order.id,
            'csv_file': encoded_file,
            'csv_delimiter': ',',
            'ignore_errors': False
        })
        with self.assertRaises(UserError):
            wizard.button_import_csv()

    def test_import_csv_with_ignore_errors(self):
        """Test import with 'ignore errors' enabled"""
        excel_file = self._create_test_excel_file()
        # Modify the test data to include an invalid quantity (non-numeric)
        excel_file.seek(0)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["product_ref", "quantity", "discount", "warehouse"])
        sheet.append(["TEST_PRODUCT_2",3, 10, "Test Warehouse"])
        sheet.append(["TEST_PRODUCT_1","INVALID_QUANTITY", 5, "Test Warehouse"])
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)
        encoded_file = base64.b64encode(excel_stream.getvalue())
        # Create the wizard with the encoded file
        wizard = self.env['sale.order.line.import'].create({
            'order_id': self.sale_order.id,
            'csv_file': encoded_file,
            'csv_delimiter': ',',
            'ignore_errors': True,  # Ignore errors
        })
        # Call the import method
        wizard.button_import_csv()
        # Verify that only valid lines were processed
        order_lines = self.sale_order.order_line
        self.assertEqual(len(order_lines), 2, "Expected 2 sale order lines after import with ignored errors")
        # Verify the valid product was imported
        product_line2 = order_lines.filtered(lambda l: l.product_id.default_code == "TEST_PRODUCT_2")
        self.assertTrue(product_line2, "Product TEST_PRODUCT_2 was not imported")
        self.assertEqual(product_line2.product_uom_qty, 3, "Incorrect quantity for TEST_PRODUCT_2")
