# -*- coding: utf-8 -*-
import base64
from io import BytesIO

from odoo.exceptions import UserError
from odoo.exceptions import ValidationError
from openpyxl import load_workbook

from odoo import models, fields

COL_CODE = "default code"
COL_NAME = "product name"
COL_VARN = "variant name"
COL_QTY = "quantity"
COL_PRICE = "price"


class PurchaseOrderLineImportCSV(models.TransientModel):
    _name = "purchase.order.line.import"

    ##########################################################################################################
    # Fields
    ##########################################################################################################
    purchase = fields.Many2one("purchase.order", string="Order Reference", readonly=True)
    file = fields.Binary("File", attachment=False, help="Select the XLS file to import")
    state = fields.Selection([("draft", "Draft"), ("done", "Done")], string="State", readonly=True, default="draft")

    ##########################################################################################################
    # Model methods
    ##########################################################################################################

    def from_template_to_product(self, product_template, variant, default_code):

        if len(product_template.product_variant_ids) == 1:
            return product_template.product_variant_id

        if not variant:
            raise ValidationError(f"Product {default_code} has variants but no variant is in the import file")

        product = False
        if product_template.product_variant_ids and variant:
            valid_variants = self.env['product.attribute.value'].search([('name', '=', variant)])
            if not valid_variants:
                raise ValidationError(
                    'Product {product} has no variant matching what has been sepecified - check caps'.format(
                        product=default_code
                    ))
            ptav = self.env['product.template.attribute.value'].search(
                [
                    ('product_tmpl_id', '=', product_template.id),
                    ('product_attribute_value_id', 'in', [x.id for x in valid_variants])
                ])
            product = self.env['product.product'].search([('product_template_attribute_value_ids', 'in', ptav)])
            if not product:
                raise ValidationError('Product {product} variant could not be resolved'.format(
                    product=default_code))
            elif len(product) > 1:
                raise ValidationError('Product {product} more than one valid variant found'.format(
                    product=default_code))
            else:
                product = product[0]
        return product

    def process_line(self, row, col_indices):
        # if not all(cell is None for cell in row):
        # if not all(cell.value is None for cell in row):
        # Just in case "default code" is non-string value
        cell = row[col_indices[COL_CODE]]
        default_code = cell.value
        if default_code == None:
            # stop when reached at empty row
            return False
        else:
            if cell.data_type != "s":
                default_code = str(default_code)

            product_name = row[col_indices[COL_NAME]].value
            variant = row[col_indices[COL_VARN]].value
            quantity = row[col_indices[COL_QTY]].value
            price = row[col_indices[COL_PRICE]].value

            product = False
            if default_code:
                # product_template = self.env['product.template'].search([('default_code','=', default_code)])
                self.env.cr.execute("SELECT id FROM product_template WHERE default_code = %s", (default_code,))
                product_ids = self.env.cr.fetchall()
                if product_ids and product_ids[0] and product_ids[0][0]:
                    product_template = self.env['product.template'].browse(product_ids[0][0])
                else:
                    product_template = None
                if product_template:
                    product = self.from_template_to_product(product_template, variant, default_code)

            elif product_name and not product:
                product_template = self.env['product.template'].search([('name', '=', product_name)])
                product = self.from_template_to_product(product_template, variant, default_code)

            if not product:
                raise ValidationError(f"Product {default_code} cannot be found")

            self.create_purchase_order_line(product, quantity, price)
        return True

    def button_process_import(self):
        """
        Process the import.
        """
        self.ensure_one()
        if self.purchase.state != "draft":
            raise UserError("Purchase Line import only available for Draft state Purchase Orders")

        file = BytesIO(base64.b64decode(self.file))
        wb = load_workbook(file, read_only=True)
        sheet = wb.active

        # Validate the input file
        colnames = (COL_CODE, COL_NAME, COL_VARN, COL_QTY, COL_PRICE)
        col_indices = {}
        for n, cell in enumerate(sheet["1"]):
            if cell.data_type != "s":
                continue
            name = cell.value.strip()
            if name in colnames:
                col_indices[name] = n
        if len(col_indices) != len(colnames):
            raise UserError("Spreadsheet does not contain header line with required column names")
        for row in range(2, sheet.max_row + 1):
            line = self.process_line(sheet[f"{row}"], col_indices)
            if not line:
                break
        return {"type": "ir.actions.act_window_close"}

    def create_purchase_order_line(self, product, quantity, price):

        line_model = self.env["purchase.order.line"]

        product_uom = product.uom_id.id
        if not price:
            supplier_recs = self.env['product.supplierinfo'].search([('product_id', '=', product.id),
                                                                     ('partner_id', '=', self.purchase.partner_id.id),
                                                                     (
                                                                         'currency_id', '=',
                                                                         self.purchase.currency_id.id)])
            if supplier_recs:
                price = supplier_recs[0].price
            else:
                price = product.standard_price
        purchase_values = {
            "product_id": product.id,
            "price_unit": price or 0,
            "order_id": self.purchase.id,
            "date_planned": fields.Date.today(),
            "product_uom_id": product_uom,
            "product_qty": quantity,
            "name": product.display_name,
        }

        record = line_model.create([purchase_values])
        record._compute_tax_id()
