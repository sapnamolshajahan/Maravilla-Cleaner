# -*- coding: utf-8 -*-

from odoo.tests import common, tagged
from odoo.exceptions import ValidationError, UserError
import base64
from openpyxl import Workbook
from io import BytesIO

from odoo.addons.sale_subscription.models.product import product_template

COL_CODE = "default code"
COL_NAME = "product name"
COL_VARN = "variant name"
COL_QTY = "quantity"
COL_PRICE = "price"

@tagged("common", "purchase_order_line_import_csv")
class TestPurchaseOrderLineImportCSV(common.TransactionCase):

    def setUp(self):
        super().setUp()

        # Setup required data
        self.currency = self.env.ref("base.USD")
        self.supplier = self.env["res.partner"].create({"name": "Test Supplier"})
        self.product1 = self.env.ref('product.product_product_6')
        self.product2 = self.env.ref('product.product_product_7')
        self.attribute = self.env['product.attribute'].create({'name': 'Size'})
        self.value_1 = self.env['product.attribute.value'].create({'name': 'Variant 1', 'attribute_id': self.attribute.id})
        self.value_2 = self.env['product.attribute.value'].create({'name': 'Variant 1', 'attribute_id': self.attribute.id})
        self.product1.write({
            "name": "Product 1",
            "default_code": "PROD1",
            "standard_price": 10.0,
            'attribute_line_ids': [(0, 0, {
                'attribute_id': self.attribute.id,
                'value_ids': [(4, self.value_1.id)]
            })]
        })
        self.product2.write({
            "name": "Product 2",
            "default_code": "PROD2",
            "standard_price": 20.0,
            'attribute_line_ids': [(0, 0, {
                'attribute_id': self.attribute.id,
                'value_ids': [(4, self.value_2.id)]
            })]
        })
        self.purchase_order = self.env["purchase.order"].create({
            "partner_id": self.supplier.id,
            "currency_id": self.currency.id,
        })

        # Create a sample XLS file for testing
        self.file_data = self._generate_sample_xls_file()

        self.wizard = self.env["purchase.order.line.import"].create({
            "purchase": self.purchase_order.id,
            "file": self.file_data,
        })

    def _generate_sample_xls_file(self):
        """Generate a sample XLS file for testing."""
        wb = Workbook()
        sheet = wb.active
        # Add headers
        sheet.append([COL_CODE, COL_NAME, COL_VARN, COL_QTY, COL_PRICE])
        # Add sample data
        sheet.append(["PROD1", "Product 1", "Variant 1", 5, 15.0])  # Line for product 1
        sheet.append(["PROD2", "Product 2", "Variant 2", 10, 25.0])  # Line for product 2

        # Save to bytes
        xls_data = BytesIO()
        wb.save(xls_data)
        xls_data.seek(0)
        return base64.b64encode(xls_data.read())

    def test_button_process_import(self):
        """Test the import process."""
        # Ensure purchase order starts in draft state
        self.assertEqual(self.purchase_order.state, "draft", "Purchase Order must be in draft state")

        # Run the wizard's import method
        self.wizard.button_process_import()

        # Validate the purchase order lines
        self.assertEqual(len(self.purchase_order.order_line), 2, "Purchase Order should have two lines")
        line1 = self.purchase_order.order_line.filtered(lambda l: l.product_id == self.product1)
        line2 = self.purchase_order.order_line.filtered(lambda l: l.product_id == self.product2)

        # Validate product 1 line
        self.assertEqual(line1.product_qty, 5, "Line 1 quantity should be 5")
        self.assertEqual(line1.price_unit, 15.0, "Line 1 price should be 15.0")

        # Validate product 2 line
        self.assertEqual(line2.product_qty, 10, "Line 2 quantity should be 10")
        self.assertEqual(line2.price_unit, 25.0, "Line 2 price should be 25.0")

    def test_validation_wrong_headers(self):
        """Test validation error for missing or incorrect headers."""
        # Generate file with incorrect headers
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Wrong Header 1", "Wrong Header 2"])
        xls_data = BytesIO()
        wb.save(xls_data)
        xls_data.seek(0)
        self.wizard.file = base64.b64encode(xls_data.read())

        with self.assertRaises(UserError):
            self.wizard.button_process_import()

    def test_default_price(self):
        """Test default price assignment when no price is provided."""
        # Modify the file data to exclude price for product 1
        wb = Workbook()
        sheet = wb.active
        sheet.append([COL_CODE, COL_NAME, COL_VARN, COL_QTY, COL_PRICE])
        sheet.append(["PROD1", "Product 1", "Variant 1", 5, None])  # No price specified
        xls_data = BytesIO()
        wb.save(xls_data)
        xls_data.seek(0)
        self.wizard.file = base64.b64encode(xls_data.read())

        self.wizard.button_process_import()

        line1 = self.purchase_order.order_line.filtered(lambda l: l.product_id == self.product1)
        self.assertEqual(line1.price_unit, 10.0, "Line 1 price should default to the product's standard price")

    def test_skip_empty_rows(self):
        """Test that empty rows are skipped."""
        # Modify the file data to include an empty row
        wb = Workbook()
        sheet = wb.active
        sheet.append([COL_CODE, COL_NAME, COL_VARN, COL_QTY, COL_PRICE])
        sheet.append(["PROD1", "Product 1", "Variant 1", 5, 15.0])
        sheet.append([None, None, None, None, None])  # Empty row
        xls_data = BytesIO()
        wb.save(xls_data)
        xls_data.seek(0)
        self.wizard.file = base64.b64encode(xls_data.read())
        self.wizard.button_process_import()
        # Validate only one line is created
        self.assertEqual(len(self.purchase_order.order_line), 1, "Only one valid line should be imported")

    def test_product_not_found(self):
        """Test validation error when a product is not found."""
        # Modify the file data to include a non-existent product code
        wb = Workbook()
        sheet = wb.active
        sheet.append([COL_CODE, COL_NAME, COL_VARN, COL_QTY, COL_PRICE])
        sheet.append(["NON_EXISTENT_CODE", "Non-existent Product", "Variant 1", 5, 15.0])
        xls_data = BytesIO()
        wb.save(xls_data)
        xls_data.seek(0)
        self.wizard.file = base64.b64encode(xls_data.read())

        with self.assertRaises(ValidationError):
            self.wizard.button_process_import()

    def test_missing_variant_in_import_file(self):
        """Test error when product has multiple variants but no variant is specified."""
        product_template = self.env.ref('product.product_product_11_product_template')
        product_template.write({
            'default_code': 'SVP',
            'attribute_line_ids': [(0, 0, {
                'attribute_id': self.attribute.id,
                'value_ids': [(4, self.value_1.id), (4, self.value_2.id)]
            })]
        })
        with self.assertRaises(ValidationError, msg="Product SVP has variants but no variant is in the import file"):
            self.wizard.from_template_to_product(
                product_template, variant=None, default_code="SVP"
            )
        with self.assertRaises(ValidationError,
                               msg="Product SVP has no variant matching what has been specified - check caps"):
            self.wizard.from_template_to_product(
                product_template, variant="Non-existent Variant", default_code="SVP"
            )
        with self.assertRaises(ValidationError,
                               msg="Product PROD1 variant could not be resolved"):
            self.wizard.from_template_to_product(
                product_template.product_variant_id, variant="Variant 1", default_code="PROD1"
            )


    def test_non_draft_purchase_order(self):
        """Test error when importing to non-draft purchase order."""
        # Change the purchase order state to 'purchase'
        self.purchase_order.state = "purchase"
        with self.assertRaises(UserError):
            self.wizard.button_process_import()
