# -*- coding: utf-8 -*-
import logging
import base64
from io import BytesIO
from openpyxl import load_workbook
from odoo.tests import common, tagged

_logger = logging.getLogger(__name__)


@tagged("common", "account_unallocated_cash_report")
class TestUnallocatedCashReport(common.TransactionCase):
    """Class to test unallocated cash report workflow"""

    def setUp(self):
        """
        Extend setUp functionality to create data used in tests
        """
        super().setUp()
        self.product = self.env.ref('product.product_product_6')
        self.partner = self.env.ref('base.res_partner_12')
        sequence = self.env['ir.sequence'].create({
            'name': 'Sales Journal Sequence',
            'implementation': 'no_gap',
            'prefix': 'INV/',
            'padding': 5,
        })
        bank_journal = self.env['account.journal'].create({
            'name': 'Bank',
            'code': 'BNKKK',
            'type': 'bank',
            'sequence_id': sequence.id
        })
        account_other = self.env['account.account'].create({
            'account_type': 'asset_receivable',
            'name': 'account_other',
            'code': '121040',

        })
        self.account_move_1 = self.env['account.move'].create({
            'partner_id': self.partner.id,
            'move_type': 'entry',
            'date': '1990-01-01',
            'journal_id': bank_journal.id,
            'line_ids': [
                (0, 0, {'name':'Test','debit': 200.0, 'credit': 0.0, 'account_id': account_other.id,
                       }),
                (0, 0, {'name':'Test','debit': 0.0, 'credit': 200.0, 'account_id': account_other.id}),
            ],
        })
        self.account_move_1.action_post()
        self.report = self.env["unallocated.cash.report"].create({})

    def test_button_process(self):
        """Test report generation"""
        self.report.button_process()
        # Check report state
        self.assertEqual(self.report.state, "done", "Report state should be 'done' after processing.")
        self.assertTrue(self.report.data, "Report data should not be empty after processing.")
        report_data = base64.b64decode(self.report.data)
        data_stream = BytesIO(report_data)
        workbook = load_workbook(data_stream)
        worksheet = workbook.active
        # Check if the first row contains the expected headers
        expected_headers = ["Type", "Customer", "Customer Number", "Invoice Number", "Line Total", "Line Balance"]
        actual_headers = [worksheet.cell(row=1, column=i).value for i in range(1, 7)]
        self.assertEqual(expected_headers, actual_headers, "Excel headers do not match expected values.")
        # Ensure at least one data row is present
        self.assertGreater(worksheet.max_row, 1, "Excel file should have data rows.")
        data_rows = [
            {
                "Type": worksheet.cell(row=r, column=1).value,
                "Customer": worksheet.cell(row=r, column=2).value,
                "Customer Number": worksheet.cell(row=r, column=3).value,
                "Invoice Number": worksheet.cell(row=r, column=4).value,
                "Line Total": worksheet.cell(row=r, column=5).value,
                "Line Balance": worksheet.cell(row=r, column=6).value,
            }
            for r in range(2, worksheet.max_row + 1)
        ]
        # Find the specific row that belongs to `self.account_move_1`
        expected_invoice_number = self.account_move_1.line_ids.filtered(lambda l: l.credit > 0).name
        matching_row = next((row for row in data_rows if row["Invoice Number"] == expected_invoice_number), None)
        # Ensure we found the correct row
        self.assertIsNotNone(matching_row, "Expected row with specific Invoice Number was not found in report.")
        # Expected values from `self.account_move_1`
        expected_values = {
            "Type": "credit-note",
            "Customer": self.partner.name,
            "Customer Number": self.partner.ref or None,
            "Invoice Number": expected_invoice_number,
            "Line Total": 200,  # From `credit` field
            "Line Balance": -200,  # From `balance` field
        }
        self.assertEqual(matching_row, expected_values, "Row data does not match expected values.")
