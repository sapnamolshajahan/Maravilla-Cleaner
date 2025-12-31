# -*- coding: utf-8 -*-
import base64
import logging
from io import BytesIO

from openpyxl import load_workbook

from odoo import models, fields

_logger = logging.getLogger(__name__)


class ImportTransferCsv(models.TransientModel):
    """
    Create transfers from a Excel
    """
    _name = "import.transfers.csv"

    picking = fields.Many2one("stock.picking", "Picking to transfer against", required=True, readonly=True)
    source_location_id = fields.Many2one("stock.location", string="Default Source Location")
    dest_location_id = fields.Many2one("stock.location", string="Default Destination Location")
    delimiter = fields.Char("Field Delimiter", size=5, required=True, default=",")
    file = fields.Binary("File", help="Select the Excel file to import.")
    notes = fields.Text("Notes", readonly=True)

    def button_import(self):
        self.picking.write(
            {
                "location_id": self.source_location_id.id,
                "location_dest_id": self.dest_location_id.id,
            })

        move_model = self.env["stock.move"]
        product_model = self.env["product.product"]

        error_count = 0

        wb = load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)
        ws = wb.active
        for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
            products = product_model.search([("default_code", '=', record[0])])
            product = products.filtered(lambda p: p.default_code == record[0])
            if len(product) > 1:
                product = product[0]
            if not product:
                _logger.warning("couldn't find product={0}".format(record[0]))
                error_count += 1
                continue
            try:
                qty = float(record[1])
            except (TypeError, ValueError):
                _logger.warning("unconvertible quantity={0}".format(record[1]))
                error_count += 1
                continue
            values = {
                "display_name": product.name,
                "description_picking": product.name,
                "location_id": self.source_location_id.id,
                "location_dest_id": self.dest_location_id.id,
                "picking_id": self.picking.id,
                "picking_type_id": self.picking.picking_type_id.id,
                "product_id": product.id,
                "product_uom": product.uom_id.id,
                "product_uom_qty": qty,
            }
            move_model.create(values)
        return {"type": "ir.actions.act_window_close"}
